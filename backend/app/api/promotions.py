"""Promotion endpoints."""

import csv
import io
from datetime import UTC, date, datetime

from fastapi import APIRouter, Depends, Form, HTTPException, UploadFile
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.auth import get_current_user
from app.core.database import get_db
from app.models.product import Product, WBAccount
from app.models.promotion import Promotion, PromotionProduct
from app.models.user import User
from app.schemas.promotion import (
    DecisionUpdate,
    PromotionDetailResponse,
    PromotionListResponse,
    PromotionProductResponse,
    PromotionResponse,
)
from app.services.promotion_collector import calculate_promo_margin, _determine_status

router = APIRouter(prefix="/promotions", tags=["promotions"])


@router.get("", response_model=PromotionListResponse)
async def list_promotions(
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """List promotions with aggregated margin stats."""
    query = select(Promotion).order_by(Promotion.start_date.desc())
    if status:
        query = query.where(Promotion.status == status)

    result = await db.execute(query)
    promotions = list(result.scalars().all())

    items = []
    for promo in promotions:
        # Aggregate promotion products stats
        agg_result = await db.execute(
            select(
                func.count(PromotionProduct.id),
                func.avg(PromotionProduct.current_margin_pct),
                func.avg(PromotionProduct.promo_margin_pct),
                func.count().filter(PromotionProduct.promo_margin_pct > 0),
            ).where(PromotionProduct.promotion_id == promo.id)
        )
        row = agg_result.one()
        products_count = row[0] or 0
        avg_current = round(float(row[1]), 1) if row[1] is not None else None
        avg_promo = round(float(row[2]), 1) if row[2] is not None else None
        profitable = row[3] or 0

        items.append(PromotionResponse(
            id=promo.id,
            wb_promo_id=promo.wb_promo_id,
            name=promo.name,
            start_date=promo.start_date,
            end_date=promo.end_date,
            promo_type=promo.promo_type,
            status=promo.status,
            in_action_count=promo.in_action_count or 0,
            total_available=promo.total_available or 0,
            products_count=products_count,
            avg_current_margin=avg_current,
            avg_promo_margin=avg_promo,
            profitable_count=profitable,
        ))

    return PromotionListResponse(items=items, total=len(items))


@router.get("/{promo_id}", response_model=PromotionDetailResponse)
async def get_promotion_detail(
    promo_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Get promotion details with all products and margins."""
    result = await db.execute(
        select(Promotion).where(Promotion.id == promo_id)
    )
    promo = result.scalar_one_or_none()
    if not promo:
        raise HTTPException(status_code=404, detail="Promotion not found")

    # Get promotion products
    pp_result = await db.execute(
        select(PromotionProduct)
        .where(PromotionProduct.promotion_id == promo_id)
        .order_by(PromotionProduct.promo_margin_pct.desc().nulls_last())
    )
    promo_products = list(pp_result.scalars().all())

    # Enrich with product info (vendor_code, title, image_url)
    nm_ids = [pp.nm_id for pp in promo_products]
    product_info = {}
    if nm_ids:
        prod_result = await db.execute(
            select(Product.nm_id, Product.vendor_code, Product.title, Product.image_url)
            .where(Product.nm_id.in_(nm_ids))
        )
        for row in prod_result.all():
            product_info[row[0]] = {
                "vendor_code": row[1],
                "title": row[2],
                "image_url": row[3],
            }

    # Aggregate stats
    agg_result = await db.execute(
        select(
            func.count(PromotionProduct.id),
            func.avg(PromotionProduct.current_margin_pct),
            func.avg(PromotionProduct.promo_margin_pct),
            func.count().filter(PromotionProduct.promo_margin_pct > 0),
        ).where(PromotionProduct.promotion_id == promo_id)
    )
    agg = agg_result.one()

    promo_response = PromotionResponse(
        id=promo.id,
        wb_promo_id=promo.wb_promo_id,
        name=promo.name,
        start_date=promo.start_date,
        end_date=promo.end_date,
        promo_type=promo.promo_type,
        status=promo.status,
        in_action_count=promo.in_action_count or 0,
        total_available=promo.total_available or 0,
        products_count=agg[0] or 0,
        avg_current_margin=round(float(agg[1]), 1) if agg[1] is not None else None,
        avg_promo_margin=round(float(agg[2]), 1) if agg[2] is not None else None,
        profitable_count=agg[3] or 0,
    )

    products = []
    for pp in promo_products:
        info = product_info.get(pp.nm_id, {})
        products.append(PromotionProductResponse(
            nm_id=pp.nm_id,
            vendor_code=info.get("vendor_code"),
            title=info.get("title"),
            image_url=info.get("image_url"),
            plan_price=float(pp.plan_price) if pp.plan_price else None,
            plan_discount=float(pp.plan_discount) if pp.plan_discount else None,
            current_price=float(pp.current_price) if pp.current_price else None,
            in_action=pp.in_action or False,
            current_margin_pct=float(pp.current_margin_pct) if pp.current_margin_pct is not None else None,
            current_margin_rub=float(pp.current_margin_rub) if pp.current_margin_rub is not None else None,
            promo_margin_pct=float(pp.promo_margin_pct) if pp.promo_margin_pct is not None else None,
            promo_margin_rub=float(pp.promo_margin_rub) if pp.promo_margin_rub is not None else None,
            decision=pp.decision or "pending",
        ))

    return PromotionDetailResponse(promotion=promo_response, products=products)


@router.post("/{promo_id}/decisions")
async def update_decisions(
    promo_id: int,
    data: DecisionUpdate,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Batch update decisions (enter/skip) for promotion products."""
    result = await db.execute(
        select(Promotion).where(Promotion.id == promo_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Promotion not found")

    if data.decision not in ("enter", "skip", "pending"):
        raise HTTPException(status_code=400, detail="Invalid decision. Use 'enter', 'skip', or 'pending'.")

    now = datetime.now(UTC)
    pp_result = await db.execute(
        select(PromotionProduct).where(
            PromotionProduct.promotion_id == promo_id,
            PromotionProduct.nm_id.in_(data.nm_ids),
        )
    )
    updated = 0
    for pp in pp_result.scalars().all():
        pp.decision = data.decision
        pp.decided_at = now
        updated += 1

    await db.commit()
    return {"updated": updated}


@router.post("/import")
async def import_promotion(
    file: UploadFile,
    name: str = Form(...),
    start_date: str = Form(""),
    end_date: str = Form(""),
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Import promotion from Excel/CSV file downloaded from WB.

    Expected columns (Russian or English, flexible matching):
    - Артикул WB / nmID / nm_id / Артикул
    - Плановая цена / planPrice / Акционная цена / План цена
    - Плановая скидка / planDiscount / Скидка акции %
    - Текущая цена / currentPrice
    - Участвует / inAction / В акции (да/нет/true/false)

    Form fields:
    - name: promotion name (required)
    - start_date: YYYY-MM-DD (optional)
    - end_date: YYYY-MM-DD (optional)
    """
    # Parse dates
    parsed_start: date | None = None
    parsed_end: date | None = None
    try:
        if start_date:
            parsed_start = date.fromisoformat(start_date)
        if end_date:
            parsed_end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    # Read file
    filename = file.filename or ""
    content = await file.read()

    rows: list[dict[str, str]] = []

    if filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: str(v) if v is not None else "" for i, v in enumerate(row) if i < len(headers)}
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=400, detail="Excel support requires openpyxl. Use CSV format.")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")
    else:
        # CSV
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1251")

        first_line = text.split("\n")[0]
        delimiter = ";" if ";" in first_line else ("," if "," in first_line else "\t")

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        for row in reader:
            rows.append({k.strip().lower(): v.strip() for k, v in row.items() if k})

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows")

    # Handle WB special case: file with only "Товар уже участвует в акции" header
    first_keys = list(rows[0].keys()) if rows else []
    if first_keys and any("уже участвует" in k for k in first_keys):
        raise HTTPException(
            status_code=400,
            detail="Все товары уже участвуют в акции. Файл от WB не содержит данных для импорта.",
        )

    # Column aliases → internal field names
    # Covers official WB Calendar export + common variations
    aliases = {
        # nm_id
        "артикул wb": "nm_id", "артикул вб": "nm_id", "артикул": "nm_id",
        "nmid": "nm_id", "nm_id": "nm_id", "nm id": "nm_id",
        "код номенклатуры": "nm_id", "номенклатура": "nm_id",
        "артикул wildberries": "nm_id",
        # plan_price (WB: "Плановая цена" — red highlight)
        "плановая цена": "plan_price", "план цена": "plan_price",
        "planprice": "plan_price", "plan_price": "plan_price",
        "акционная цена": "plan_price", "цена акции": "plan_price",
        "цена для акции": "plan_price", "загруженная цена": "plan_price",
        "плановая акционная цена": "plan_price",
        # plan_discount (WB: "Загружаемая скидка" — green highlight)
        "плановая скидка": "plan_discount", "скидка акции": "plan_discount",
        "скидка акции %": "plan_discount", "plandiscount": "plan_discount",
        "plan_discount": "plan_discount", "скидка для акции": "plan_discount",
        "загружаемая скидка": "plan_discount",
        # current_price
        "текущая цена": "current_price", "currentprice": "current_price",
        "current_price": "current_price", "цена до скидки": "current_price",
        "цена": "current_price", "текущая розничная цена": "current_price",
        "текущая цена до скидки": "current_price",
        # in_action (WB: "Участие в акции" — Да/Нет)
        "участвует": "in_action", "в акции": "in_action",
        "inaction": "in_action", "in_action": "in_action",
        "статус участия": "in_action", "участие в акции": "in_action",
    }

    normalized_rows = []
    for row in rows:
        norm: dict[str, str] = {}
        for k, v in row.items():
            key = aliases.get(k, k)
            norm[key] = v
        normalized_rows.append(norm)

    # Get first account
    acc_result = await db.execute(select(WBAccount).where(WBAccount.is_active == True).limit(1))  # noqa: E712
    account = acc_result.scalar_one_or_none()
    if not account:
        raise HTTPException(status_code=400, detail="No active WB account")

    account_id = account.id
    account_tax = float(account.tax_rate) if account.tax_rate else 0.0
    account_tariff = float(account.tariff_rate) if account.tariff_rate else 0.0

    # Create promotion
    status = _determine_status(parsed_start, parsed_end)
    promo = Promotion(
        account_id=account_id,
        name=name,
        start_date=parsed_start,
        end_date=parsed_end,
        promo_type="import",
        status=status,
        is_active=status != "ended",
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )
    db.add(promo)
    await db.flush()  # get promo.id

    # Load products for margin calculation
    prod_result = await db.execute(
        select(Product).where(Product.account_id == account_id)
    )
    products = list(prod_result.scalars().all())
    nm_to_product = {p.nm_id: p for p in products}

    imported = 0
    errors: list[str] = []

    for i, row in enumerate(normalized_rows, start=2):
        nm_id_str = row.get("nm_id", "")
        if not nm_id_str:
            continue

        try:
            nm_id = int(float(nm_id_str))
        except (ValueError, TypeError):
            errors.append(f"Row {i}: invalid nm_id '{nm_id_str}'")
            continue

        # Parse plan_price
        plan_price_str = row.get("plan_price", "")
        plan_price = None
        if plan_price_str:
            try:
                plan_price = float(plan_price_str.replace(",", ".").replace(" ", ""))
            except ValueError:
                pass

        # Parse plan_discount
        plan_discount_str = row.get("plan_discount", "")
        plan_discount = None
        if plan_discount_str:
            try:
                plan_discount = float(plan_discount_str.replace(",", ".").replace("%", "").replace(" ", ""))
            except ValueError:
                pass

        # Parse current_price
        current_price_str = row.get("current_price", "")
        current_price = None
        if current_price_str:
            try:
                current_price = float(current_price_str.replace(",", ".").replace(" ", ""))
            except ValueError:
                pass

        # Parse in_action
        in_action_str = row.get("in_action", "").lower()
        in_action = in_action_str in ("да", "yes", "true", "1", "участвует")

        # Calculate margins
        product = nm_to_product.get(nm_id)
        current_margin_pct = None
        current_margin_rub = None
        promo_margin_pct = None
        promo_margin_rub = None

        if product:
            if current_price:
                current_margin_pct, current_margin_rub = calculate_promo_margin(
                    product, current_price, account_tax, account_tariff
                )
            if plan_price:
                promo_margin_pct, promo_margin_rub = calculate_promo_margin(
                    product, plan_price, account_tax, account_tariff
                )

        db.add(PromotionProduct(
            promotion_id=promo.id,
            account_id=account_id,
            nm_id=nm_id,
            plan_price=plan_price,
            plan_discount=plan_discount,
            current_price=current_price,
            in_action=in_action,
            promo_price=plan_price,
            current_margin_pct=current_margin_pct,
            current_margin_rub=current_margin_rub,
            promo_margin_pct=promo_margin_pct,
            promo_margin_rub=promo_margin_rub,
            decision="pending",
        ))
        imported += 1

    # Update promotion counts
    promo.total_available = imported
    promo.in_action_count = sum(1 for r in normalized_rows if r.get("in_action", "").lower() in ("да", "yes", "true", "1", "участвует"))

    await db.commit()

    return {
        "promotion_id": promo.id,
        "name": promo.name,
        "imported": imported,
        "errors": errors[:50],
    }
