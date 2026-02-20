"""Product endpoints."""

import csv
import io
from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.auth import get_current_user
from app.core.database import get_db
from app.models.price import PriceSnapshot
from app.models.product import Product, WBAccount
from app.models.sales import SalesDaily
from app.models.user import User
from app.schemas.product import ProductCostUpdate, ProductList, ProductResponse

router = APIRouter(prefix="/products", tags=["products"])


async def _enrich_with_prices(
    db: AsyncSession, products: list[Product]
) -> list[ProductResponse]:
    """Attach latest price snapshot, orders_7d, and margin to each product."""
    if not products:
        return []

    product_ids = [p.id for p in products]

    # Get latest snapshot per product using DISTINCT ON
    latest_prices_q = (
        select(PriceSnapshot)
        .where(PriceSnapshot.product_id.in_(product_ids))
        .distinct(PriceSnapshot.product_id)
        .order_by(PriceSnapshot.product_id, PriceSnapshot.collected_at.desc())
    )
    result = await db.execute(latest_prices_q)
    snapshots = {s.product_id: s for s in result.scalars().all()}

    # Get orders and returns for last 7 full days (excluding today), Moscow timezone
    MSK = timezone(timedelta(hours=3))
    today_msk = datetime.now(MSK).date()
    seven_days_ago = today_msk - timedelta(days=7)
    orders_q = (
        select(
            SalesDaily.product_id,
            func.sum(SalesDaily.orders_count),
            func.sum(SalesDaily.returns_count),
        )
        .where(
            SalesDaily.product_id.in_(product_ids),
            SalesDaily.date >= seven_days_ago,
            SalesDaily.date < today_msk,
        )
        .group_by(SalesDaily.product_id)
    )
    orders_result = await db.execute(orders_q)
    orders_map = {
        row[0]: max((row[1] or 0) - (row[2] or 0), 0)
        for row in orders_result.all()
    }

    # Load account tax_rate and tariff_rate
    account_ids = {p.account_id for p in products}
    account_settings: dict[int, tuple[float, float]] = {}
    if account_ids:
        acc_result = await db.execute(
            select(WBAccount.id, WBAccount.tax_rate, WBAccount.tariff_rate)
            .where(WBAccount.id.in_(account_ids))
        )
        for row in acc_result.all():
            tax = float(row[1]) if row[1] else 0.0
            tariff = float(row[2]) if row[2] else 0.0
            account_settings[row[0]] = (tax, tariff)

    items = []
    for p in products:
        snap = snapshots.get(p.id)
        final_price = float(snap.final_price) if snap and snap.final_price else None
        cost_price = float(p.cost_price) if p.cost_price is not None else None
        commission_pct = float(p.commission_pct) if p.commission_pct is not None else None
        logistics_cost = float(p.logistics_cost) if p.logistics_cost is not None else None
        storage_cost = float(p.storage_cost) if p.storage_cost is not None else None
        storage_daily = float(p.storage_daily) if p.storage_daily is not None else None
        ad_pct = float(p.ad_pct) if p.ad_pct is not None else None

        # Get account-level settings
        acc_tax, acc_tariff = account_settings.get(p.account_id, (0.0, 0.0))

        # Calculate margin with all costs:
        # margin = final_price - cost_price - tax - commission - tariff - logistics - storage - ad(%)
        margin_pct = None
        margin_rub = None
        if final_price and cost_price and final_price > 0:
            tax_amount = final_price * acc_tax / 100 if acc_tax else 0
            commission_amount = final_price * commission_pct / 100 if commission_pct else 0
            tariff_amount = final_price * acc_tariff / 100 if acc_tariff else 0
            ad_amount = final_price * ad_pct / 100 if ad_pct else 0
            total_costs = (
                cost_price
                + tax_amount
                + commission_amount
                + tariff_amount
                + (logistics_cost or 0)
                + (storage_cost or 0)
                + ad_amount
            )
            margin_rub = round(final_price - total_costs, 2)
            margin_pct = round(margin_rub / final_price * 100, 1)

        items.append(
            ProductResponse(
                id=p.id,
                nm_id=p.nm_id,
                vendor_code=p.vendor_code,
                brand=p.brand,
                category=p.category,
                title=p.title,
                image_url=p.image_url,
                cost_price=cost_price,
                total_stock=p.total_stock,
                is_active=p.is_active,
                is_locomotive=p.is_locomotive,
                created_at=p.created_at,
                current_price=float(snap.wb_price) if snap else None,
                discount_pct=float(snap.wb_discount) if snap else None,
                final_price=final_price,
                commission_pct=commission_pct,
                logistics_cost=logistics_cost,
                storage_cost=storage_cost,
                storage_daily=storage_daily,
                ad_pct=ad_pct,
                orders_7d=orders_map.get(p.id, 0),
                margin_pct=margin_pct,
                margin_rub=margin_rub,
            )
        )
    return items


@router.get("/cost-template")
async def download_cost_template(
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Download CSV template with current products for cost import."""
    result = await db.execute(
        select(Product).order_by(Product.nm_id)
    )
    products = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Артикул WB", "Артикул", "Название", "Себестоимость", "Реклама %"])
    for p in products:
        writer.writerow([
            p.nm_id,
            p.vendor_code or "",
            p.title or "",
            float(p.cost_price) if p.cost_price else "",
            float(p.ad_pct) if p.ad_pct else "",
        ])

    # Encode with UTF-8 BOM for correct display in Excel on Windows
    csv_bytes = ("\ufeff" + output.getvalue()).encode("utf-8")
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=cost_template.csv"},
    )


@router.post("/import-costs")
async def import_costs(
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Import cost prices from CSV or Excel file.

    CSV format: nm_id;cost_price;ad_pct (semicolon-separated, header row required).
    Excel: .xlsx with same columns.
    """
    filename = file.filename or ""
    content = await file.read()

    rows: list[dict[str, str]] = []

    if filename.endswith((".xlsx", ".xls")):
        # Excel import
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: str(v) if v is not None else "" for i, v in enumerate(row) if i < len(headers)}
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=400, detail="Excel support requires openpyxl. Use CSV format.")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")
    else:
        # CSV import — try semicolon first, then comma
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1251")

        # Detect delimiter
        first_line = text.split("\n")[0]
        delimiter = ";" if ";" in first_line else ","

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        for row in reader:
            # Normalize keys to lowercase
            rows.append({k.strip().lower(): v.strip() for k, v in row.items() if k})

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows")

    # Map Russian headers to internal field names
    header_aliases = {
        "артикул wb": "nm_id", "артикул вб": "nm_id",
        "себестоимость": "cost_price", "себест.": "cost_price", "закупочная": "cost_price",
        "реклама %": "ad_pct", "реклама": "ad_pct", "рекл.": "ad_pct",
        "реклама ₽": "ad_pct",  # backwards compat alias
    }
    normalized_rows = []
    for row in rows:
        norm = {}
        for k, v in row.items():
            key = header_aliases.get(k, k)  # k is already lowercased
            norm[key] = v
        normalized_rows.append(norm)
    rows = normalized_rows

    # Build nm_id lookup
    result = await db.execute(select(Product))
    all_products = {p.nm_id: p for p in result.scalars().all()}

    updated = 0
    skipped = 0
    errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        nm_id_str = row.get("nm_id", "")
        cost_str = row.get("cost_price", "")
        ad_str = row.get("ad_pct", "")

        if not nm_id_str:
            skipped += 1
            continue

        try:
            nm_id = int(float(nm_id_str))
        except (ValueError, TypeError):
            errors.append(f"Row {i}: invalid nm_id '{nm_id_str}'")
            continue

        product = all_products.get(nm_id)
        if not product:
            errors.append(f"Row {i}: nm_id {nm_id} not found")
            continue

        if cost_str:
            try:
                product.cost_price = float(cost_str.replace(",", "."))
            except ValueError:
                errors.append(f"Row {i}: invalid cost_price '{cost_str}'")
                continue

        if ad_str:
            try:
                product.ad_pct = float(ad_str.replace(",", "."))
            except ValueError:
                errors.append(f"Row {i}: invalid ad_pct '{ad_str}'")

        updated += 1

    await db.commit()
    return {"updated": updated, "skipped": skipped, "errors": errors[:50]}


@router.get("/export-costs")
async def export_costs(
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    """Export current products with costs as CSV."""
    result = await db.execute(select(Product).order_by(Product.nm_id))
    products = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Артикул WB", "Артикул", "Бренд", "Название", "Себестоимость", "Комиссия %", "Логистика ₽", "Хранение ₽/прод", "Хранение ₽/сут", "Реклама %", "Остаток"])
    for p in products:
        writer.writerow([
            p.nm_id,
            p.vendor_code or "",
            p.brand or "",
            p.title or "",
            float(p.cost_price) if p.cost_price else "",
            float(p.commission_pct) if p.commission_pct else "",
            float(p.logistics_cost) if p.logistics_cost else "",
            float(p.storage_cost) if p.storage_cost else "",
            float(p.storage_daily) if p.storage_daily else "",
            float(p.ad_pct) if p.ad_pct else "",
            p.total_stock,
        ])

    # Encode with UTF-8 BOM for correct display in Excel on Windows
    csv_bytes = ("\ufeff" + output.getvalue()).encode("utf-8")
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=products_costs.csv"},
    )


@router.get("", response_model=ProductList)
async def list_products(
    skip: int = 0,
    limit: int = 50,
    search: str | None = None,
    brand: str | None = None,
    category: str | None = None,
    is_active: bool | None = None,
    in_stock: bool | None = None,
    sort_by: str | None = None,
    sort_order: str = "ascend",
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    query = select(Product)

    if search:
        pattern = f"%{search}%"
        query = query.where(
            or_(
                Product.title.ilike(pattern),
                Product.vendor_code.ilike(pattern),
                Product.brand.ilike(pattern),
            )
        )
    if brand:
        query = query.where(Product.brand == brand)
    if category:
        query = query.where(Product.category == category)
    if is_active is not None:
        query = query.where(Product.is_active == is_active)
    if in_stock:
        query = query.where(Product.total_stock >= 1)

    # Load all matching products, enrich, sort, then paginate
    result = await db.execute(query.order_by(Product.nm_id))
    all_products = list(result.scalars().all())

    items = await _enrich_with_prices(db, all_products)

    # Server-side sorting on enriched data
    if sort_by:
        reverse = sort_order == "descend"
        items.sort(key=lambda x: getattr(x, sort_by, 0) or 0, reverse=reverse)

    total = len(items)
    page_items = items[skip : skip + limit]
    return ProductList(items=page_items, total=total)


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    enriched = await _enrich_with_prices(db, [product])
    return enriched[0]


@router.put("/{product_id}/cost", response_model=ProductResponse)
async def update_product_cost(
    product_id: int,
    data: ProductCostUpdate,
    db: AsyncSession = Depends(get_db),
    _current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    if data.cost_price is not None:
        product.cost_price = data.cost_price
    if data.ad_pct is not None:
        product.ad_pct = data.ad_pct

    enriched = await _enrich_with_prices(db, [product])
    return enriched[0]
